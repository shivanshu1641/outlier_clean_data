"""Tests for format conversion and format-specific corruptions."""

from __future__ import annotations

import io
import json
import random

import numpy as np
import pandas as pd
import pytest

from server.corruption.format_corruptions import (
    FORMAT_CORRUPTION_FNS,
    FORMAT_CORRUPTIONS,
    SUPPORTED_FORMATS,
    apply_format_corruptions,
    convert_to_format,
    format_preview,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_df() -> pd.DataFrame:
    return pd.DataFrame({
        "name": ["Alice", "Bob", "Charlie", "Diana", "Eve"],
        "age": [30, 25, 35, 28, 22],
        "score": [88.5, 92.0, 76.3, 95.1, 81.0],
        "city": ["NYC", "LA", "Chicago", "Houston", "Seattle"],
    })


@pytest.fixture
def rng():
    return np.random.default_rng(42)


@pytest.fixture
def py_rng():
    return random.Random(42)


# ---------------------------------------------------------------------------
# Part 1 -- Format Conversion Tests
# ---------------------------------------------------------------------------

class TestConvertToFormat:
    """All format conversions produce non-empty output."""

    @pytest.mark.parametrize("fmt", SUPPORTED_FORMATS)
    def test_non_empty(self, sample_df: pd.DataFrame, fmt: str):
        result = convert_to_format(sample_df, fmt)
        assert result, f"convert_to_format returned empty output for {fmt}"
        if fmt == "excel":
            assert isinstance(result, bytes)
        else:
            assert isinstance(result, str)

    def test_unsupported_format_raises(self, sample_df: pd.DataFrame):
        with pytest.raises(ValueError, match="Unsupported format"):
            convert_to_format(sample_df, "parquet_v99")

    def test_csv_roundtrip(self, sample_df: pd.DataFrame):
        csv_str = convert_to_format(sample_df, "csv")
        restored = pd.read_csv(io.StringIO(csv_str))
        pd.testing.assert_frame_equal(sample_df, restored)

    def test_tsv_roundtrip(self, sample_df: pd.DataFrame):
        tsv_str = convert_to_format(sample_df, "tsv")
        restored = pd.read_csv(io.StringIO(tsv_str), sep="\t")
        pd.testing.assert_frame_equal(sample_df, restored)

    def test_json_roundtrip(self, sample_df: pd.DataFrame):
        j = convert_to_format(sample_df, "json")
        records = json.loads(j)
        restored = pd.DataFrame(records)
        pd.testing.assert_frame_equal(sample_df, restored)

    def test_jsonl_roundtrip(self, sample_df: pd.DataFrame):
        jl = convert_to_format(sample_df, "jsonl")
        records = [json.loads(line) for line in jl.strip().split("\n")]
        restored = pd.DataFrame(records)
        pd.testing.assert_frame_equal(sample_df, restored)

    def test_excel_produces_valid_workbook(self, sample_df: pd.DataFrame):
        data = convert_to_format(sample_df, "excel")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == len(sample_df) + 1  # header + data
        assert ws.max_column == len(sample_df.columns)
        wb.close()

    def test_xml_contains_tags(self, sample_df: pd.DataFrame):
        xml = convert_to_format(sample_df, "xml")
        assert "<data>" in xml
        assert "<record>" in xml
        assert "</data>" in xml

    def test_fixed_width_has_aligned_columns(self, sample_df: pd.DataFrame):
        fw = convert_to_format(sample_df, "fixed_width")
        lines = fw.split("\n")
        # Ignore empty trailing line if present
        lines = [l for l in lines if l]
        # All lines should have the same length (after right-padding)
        max_len = max(len(line) for line in lines)
        padded_lengths = {len(line.ljust(max_len)) for line in lines}
        assert len(padded_lengths) == 1, f"Fixed width lines have inconsistent lengths: {set(len(l) for l in lines)}"

    def test_html_table_has_tags(self, sample_df: pd.DataFrame):
        html = convert_to_format(sample_df, "html_table")
        assert "<table" in html
        assert "<th>" in html or "<td>" in html

    def test_sql_dump_has_inserts(self, sample_df: pd.DataFrame):
        sql = convert_to_format(sample_df, "sql_dump")
        assert "CREATE TABLE" in sql
        assert "INSERT INTO" in sql
        assert sql.count("INSERT INTO") == len(sample_df)

    def test_yaml_has_keys(self, sample_df: pd.DataFrame):
        yml = convert_to_format(sample_df, "yaml")
        assert "name:" in yml
        assert "age:" in yml


# ---------------------------------------------------------------------------
# Part 2 -- Preview Tests
# ---------------------------------------------------------------------------

class TestFormatPreview:
    def test_text_preview(self, sample_df: pd.DataFrame):
        csv_str = convert_to_format(sample_df, "csv")
        preview = format_preview(csv_str, "csv")
        assert "name" in preview

    def test_binary_preview_excel(self, sample_df: pd.DataFrame):
        data = convert_to_format(sample_df, "excel")
        preview = format_preview(data, "excel")
        assert "[Excel workbook" in preview or "[Binary" in preview

    def test_truncation(self, sample_df: pd.DataFrame):
        csv_str = convert_to_format(sample_df, "csv")
        preview = format_preview(csv_str, "csv", max_lines=2)
        assert "more lines" in preview


# ---------------------------------------------------------------------------
# Part 3 -- Corruption Registry Tests
# ---------------------------------------------------------------------------

class TestCorruptionRegistry:
    def test_all_format_corruptions_registered(self):
        """Every corruption name in FORMAT_CORRUPTIONS must exist in FORMAT_CORRUPTION_FNS."""
        for fmt, names in FORMAT_CORRUPTIONS.items():
            for name in names:
                assert name in FORMAT_CORRUPTION_FNS, (
                    f"Corruption {name!r} (format={fmt}) is not in FORMAT_CORRUPTION_FNS"
                )

    def test_no_orphan_corruption_fns(self):
        """Every fn in FORMAT_CORRUPTION_FNS must belong to at least one format."""
        all_names = set()
        for names in FORMAT_CORRUPTIONS.values():
            all_names.update(names)
        for name in FORMAT_CORRUPTION_FNS:
            assert name in all_names, f"{name!r} in FNS but not assigned to any format"

    def test_fns_are_callable(self):
        for name, fn in FORMAT_CORRUPTION_FNS.items():
            assert callable(fn), f"{name!r} is not callable"


# ---------------------------------------------------------------------------
# Part 4 -- Corruption Smoke Tests (each text format)
# ---------------------------------------------------------------------------

class TestCorruptionSmoke:
    """Run each corruption on real content and verify it doesn't crash."""

    @pytest.mark.parametrize("fmt", [f for f in SUPPORTED_FORMATS if f != "excel"])
    def test_text_format_corruptions_no_crash(
        self, sample_df: pd.DataFrame, fmt: str, rng, py_rng
    ):
        content = convert_to_format(sample_df, fmt)
        for name in FORMAT_CORRUPTIONS.get(fmt, []):
            fn = FORMAT_CORRUPTION_FNS[name]
            result, meta = fn(content, rng, py_rng)
            assert result is not None, f"{name} returned None for {fmt}"
            assert isinstance(meta, dict), f"{name} meta should be dict, got {type(meta)}"

    def test_excel_corruptions_no_crash(self, sample_df: pd.DataFrame, rng, py_rng):
        content = convert_to_format(sample_df, "excel")
        for name in FORMAT_CORRUPTIONS["excel"]:
            fn = FORMAT_CORRUPTION_FNS[name]
            result, meta = fn(content, rng, py_rng)
            assert isinstance(result, bytes), f"{name} should return bytes for excel"
            assert isinstance(meta, dict)

    @pytest.mark.parametrize("difficulty", ["easy", "medium", "hard"])
    def test_apply_format_corruptions_difficulty(
        self, sample_df: pd.DataFrame, rng, py_rng, difficulty: str
    ):
        content = convert_to_format(sample_df, "csv")
        result, meta_list = apply_format_corruptions(
            content, "csv", rng, py_rng, difficulty=difficulty
        )
        assert isinstance(result, str)
        assert isinstance(meta_list, list)
        assert len(meta_list) >= 1

    def test_apply_format_corruptions_unknown_format(self, rng, py_rng):
        result, meta = apply_format_corruptions("data", "unknown_fmt", rng, py_rng)
        assert result == "data"
        assert meta == []

    @pytest.mark.parametrize("fmt", SUPPORTED_FORMATS)
    def test_apply_all_formats(self, sample_df: pd.DataFrame, fmt: str, rng, py_rng):
        content = convert_to_format(sample_df, fmt)
        result, meta_list = apply_format_corruptions(content, fmt, rng, py_rng, difficulty="medium")
        assert result is not None
        assert isinstance(meta_list, list)
