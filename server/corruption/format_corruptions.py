"""Format conversion and format-specific corruption functions.

Part 1 -- converters: translate a pandas DataFrame into 9+ output formats.
Part 2 -- corruptions: inject format-level noise (structural, not value-level).
"""

from __future__ import annotations

import io
import json
import re
from typing import Any

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Part 1 -- Format Converters
# ---------------------------------------------------------------------------

SUPPORTED_FORMATS = [
    "csv",
    "tsv",
    "json",
    "jsonl",
    "excel",
    "xml",
    "fixed_width",
    "html_table",
    "sql_dump",
    "yaml",
]


def convert_to_format(df: pd.DataFrame, fmt: str) -> str | bytes:
    """Convert *df* to the requested format string (or bytes for excel)."""
    fmt = fmt.lower()
    fn = _CONVERTERS.get(fmt)
    if fn is None:
        raise ValueError(f"Unsupported format: {fmt!r}. Choose from {SUPPORTED_FORMATS}")
    return fn(df)


# --- individual converters ------------------------------------------------

def _to_csv(df: pd.DataFrame) -> str:
    return df.to_csv(index=False)


def _to_tsv(df: pd.DataFrame) -> str:
    return df.to_csv(index=False, sep="\t")


def _to_json(df: pd.DataFrame) -> str:
    return df.to_json(orient="records", indent=2, default_handler=str)


def _to_jsonl(df: pd.DataFrame) -> str:
    # Performance: use built-in vectorised path, NOT iterrows
    return df.to_json(orient="records", lines=True, default_handler=str)


def _to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


def _to_xml(df: pd.DataFrame) -> str:
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', "<data>"]
    cols = list(df.columns)
    for row in df.itertuples(index=False):
        lines.append("  <record>")
        for col, val in zip(cols, row):
            safe_tag = re.sub(r"[^A-Za-z0-9_]", "_", str(col))
            val_str = "" if pd.isna(val) else str(val)
            val_str = val_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            lines.append(f"    <{safe_tag}>{val_str}</{safe_tag}>")
        lines.append("  </record>")
    lines.append("</data>")
    return "\n".join(lines) + "\n"


def _to_fixed_width(df: pd.DataFrame) -> str:
    # Vectorised: compute column widths then format all at once
    str_df = df.astype(str).fillna("")
    widths = {c: int(max(len(c), int(str_df[c].str.len().max()))) + 2 for c in str_df.columns}
    header = "".join(c.ljust(widths[c]) for c in str_df.columns)
    rows = []
    for col in str_df.columns:
        rows.append(str_df[col].str.pad(widths[col], side="right"))
    body = pd.concat(rows, axis=1)
    lines = [header] + ["".join(vals) for vals in body.values]
    return "\n".join(lines)


def _to_html_table(df: pd.DataFrame) -> str:
    return df.to_html(index=False)


def _to_sql_dump(df: pd.DataFrame) -> str:
    table = "data_table"
    cols = list(df.columns)
    col_defs = ", ".join(f'"{c}" TEXT' for c in cols)
    lines = [f"CREATE TABLE {table} ({col_defs});", ""]
    for row in df.itertuples(index=False):
        vals = []
        for v in row:
            if pd.isna(v):
                vals.append("NULL")
            elif isinstance(v, (int, float, np.integer, np.floating)):
                vals.append(str(v))
            else:
                escaped = str(v).replace("'", "''")
                vals.append(f"'{escaped}'")
        lines.append(f"INSERT INTO {table} VALUES ({', '.join(vals)});")
    return "\n".join(lines) + "\n"


def _to_yaml(df: pd.DataFrame) -> str:
    # Lightweight YAML serialiser -- avoids PyYAML dependency at runtime for
    # simple tabular data.  Produces a list-of-dicts representation.
    lines: list[str] = []
    cols = list(df.columns)
    for row in df.itertuples(index=False):
        lines.append("-")
        for col, val in zip(cols, row):
            if pd.isna(val):
                rendered = "null"
            elif isinstance(val, bool):
                rendered = "true" if val else "false"
            elif isinstance(val, (int, float, np.integer, np.floating)):
                rendered = str(val)
            else:
                rendered = f'"{str(val)}"'
            lines.append(f"  {col}: {rendered}")
    return "\n".join(lines) + "\n"


_CONVERTERS: dict[str, Any] = {
    "csv": _to_csv,
    "tsv": _to_tsv,
    "json": _to_json,
    "jsonl": _to_jsonl,
    "excel": _to_excel,
    "xml": _to_xml,
    "fixed_width": _to_fixed_width,
    "html_table": _to_html_table,
    "sql_dump": _to_sql_dump,
    "yaml": _to_yaml,
}


# ---------------------------------------------------------------------------
# format_preview -- human-readable preview of any format (incl. binary)
# ---------------------------------------------------------------------------

def format_preview(content: str | bytes, fmt: str, max_lines: int = 30) -> str:
    """Return a human-readable string preview for *content* in the given *fmt*."""
    if isinstance(content, bytes):
        if fmt == "excel":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                sheets = wb.sheetnames
                preview_lines = [f"[Excel workbook -- sheets: {', '.join(sheets)}]"]
                ws = wb[sheets[0]]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_lines:
                        preview_lines.append("...")
                        break
                    preview_lines.append("\t".join(str(c) if c is not None else "" for c in row))
                wb.close()
                return "\n".join(preview_lines)
            except Exception:
                return f"[Binary excel data, {len(content)} bytes]"
        return f"[Binary data, {len(content)} bytes]"
    lines = content.splitlines()
    if len(lines) <= max_lines:
        return content
    return "\n".join(lines[:max_lines]) + f"\n... ({len(lines) - max_lines} more lines)"


# ---------------------------------------------------------------------------
# Part 2 -- Format-Specific Corruptions
# ---------------------------------------------------------------------------

FORMAT_CORRUPTIONS: dict[str, list[str]] = {
    "csv": [
        "mixed_delimiter",
        "bom_marker",
        "mixed_encoding",
        "mixed_line_endings",
        "quoted_inconsistency",
        "header_duplicate",
    ],
    "tsv": ["tab_in_field", "mixed_tab_space"],
    "json": [
        "key_rename",
        "nested_inconsistency",
        "missing_keys",
        "mixed_value_types",
        "extra_wrapper",
        "json_syntax_noise",
        "array_irregularity",
    ],
    "jsonl": ["key_rename", "missing_keys", "mixed_value_types", "array_irregularity"],
    "excel": [
        "merged_cells",
        "multi_sheet_scatter",
        "formula_artifacts",
        "header_misalignment",
        "mixed_header_rows",
    ],
    "xml": [
        "inconsistent_tags",
        "attributes_vs_elements",
        "namespace_mess",
        "missing_closing_tags",
    ],
    "fixed_width": ["column_misalignment", "inconsistent_padding", "shifted_fields"],
    "html_table": ["colspan_rowspan_chaos", "nested_tables", "missing_td", "hidden_data"],
    "sql_dump": ["syntax_errors", "inconsistent_quoting", "mixed_null"],
    "yaml": ["indentation_errors", "mixed_tabs_spaces", "type_ambiguity"],
}

# Type alias for corruption functions
_CorrFn = Any  # Callable[[str|bytes, np.random.Generator, random.Random], tuple[str|bytes, dict]]

# ---- helpers -------------------------------------------------------------

def _pick_rows(lines: list[str], rng: np.random.Generator, frac: float = 0.2) -> list[int]:
    """Return indices of ~*frac* of rows (skip header at index 0)."""
    if len(lines) <= 1:
        return []
    data_idx = list(range(1, len(lines)))
    n = max(1, int(len(data_idx) * frac))
    return sorted(rng.choice(data_idx, size=min(n, len(data_idx)), replace=False).tolist())


def _parse_json_or_jsonl(content: str) -> tuple[list[dict], bool]:
    """Parse content as JSON array or JSONL.  Returns (records, is_jsonl)."""
    stripped = content.strip()
    if stripped.startswith("["):
        data = json.loads(content)
        return (data if isinstance(data, list) else [data]), False
    # JSONL: one JSON object per non-empty line
    records = [json.loads(line) for line in stripped.splitlines() if line.strip()]
    return records, True


def _dump_json_or_jsonl(records: list, is_jsonl: bool) -> str:
    """Serialize records back to JSON array or JSONL."""
    if is_jsonl:
        return "\n".join(json.dumps(r, default=str) for r in records) + "\n"
    return json.dumps(records, indent=2, default=str)


# ====================================================================
# CSV corruptions
# ====================================================================

def _mixed_delimiter(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Swap commas for tabs or semicolons in ~20 % of data rows."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.2)
    delims = ["\t", ";"]
    changed = 0
    for i in targets:
        d = py_rng.choice(delims)
        lines[i] = lines[i].replace(",", d)
        changed += 1
    return "\n".join(lines), {"mixed_delimiter": {"rows_affected": changed}}


def _bom_marker(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Prepend UTF-8 BOM."""
    return "\ufeff" + content, {"bom_marker": True}


def _mixed_encoding(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Insert latin-1 style characters in random fields."""
    replacements = {"a": "\xe4", "o": "\xf6", "u": "\xfc", "e": "\xe9"}
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.15)
    for i in targets:
        char_from = py_rng.choice(list(replacements.keys()))
        lines[i] = lines[i].replace(char_from, replacements[char_from], 1)
    return "\n".join(lines), {"mixed_encoding": {"rows_affected": len(targets)}}


def _mixed_line_endings(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Mix \\n, \\r\\n and \\r line endings."""
    lines = content.split("\n")
    endings = ["\n", "\r\n", "\r"]
    out: list[str] = []
    for ln in lines:
        out.append(ln + py_rng.choice(endings))
    return "".join(out), {"mixed_line_endings": True}


def _quoted_inconsistency(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Add or remove quotes inconsistently around fields."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.25)
    for i in targets:
        fields = lines[i].split(",")
        for j in range(len(fields)):
            if py_rng.random() < 0.4:
                fields[j] = f'"{fields[j]}"'
        lines[i] = ",".join(fields)
    return "\n".join(lines), {"quoted_inconsistency": {"rows_affected": len(targets)}}


def _header_duplicate(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Duplicate the header row somewhere in the data."""
    lines = content.split("\n")
    if len(lines) < 2:
        return content, {"header_duplicate": False}
    header = lines[0]
    pos = py_rng.randint(1, max(1, len(lines) - 1))
    lines.insert(pos, header)
    return "\n".join(lines), {"header_duplicate": {"inserted_at": pos}}


# ====================================================================
# TSV corruptions
# ====================================================================

def _tab_in_field(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Insert stray tab characters within field values."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.15)
    for i in targets:
        fields = lines[i].split("\t")
        if fields:
            idx = py_rng.randint(0, len(fields) - 1)
            mid = len(fields[idx]) // 2
            fields[idx] = fields[idx][:mid] + "\t" + fields[idx][mid:]
            lines[i] = "\t".join(fields)
    return "\n".join(lines), {"tab_in_field": {"rows_affected": len(targets)}}


def _mixed_tab_space(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Replace some tab delimiters with spaces."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.2)
    for i in targets:
        # Replace first tab with spaces
        lines[i] = lines[i].replace("\t", "    ", 1)
    return "\n".join(lines), {"mixed_tab_space": {"rows_affected": len(targets)}}


# ====================================================================
# JSON corruptions
# ====================================================================

def _key_rename(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Rename keys between camelCase and snake_case variants."""
    records, is_jsonl = _parse_json_or_jsonl(content)
    if not records:
        return content, {"key_rename": False}
    targets = rng.choice(len(records), size=max(1, len(records) // 5), replace=False)
    renamed: list[str] = []
    for idx in targets:
        rec = records[int(idx)]
        if not isinstance(rec, dict):
            continue
        new_rec = {}
        for k, v in rec.items():
            if "_" in k:
                # snake_case -> camelCase
                parts = k.split("_")
                new_k = parts[0] + "".join(p.capitalize() for p in parts[1:])
            else:
                # camelCase -> snake_case
                new_k = re.sub(r"([A-Z])", r"_\1", k).lower().lstrip("_")
            new_rec[new_k] = v
            if new_k != k:
                renamed.append(f"{k}->{new_k}")
        records[int(idx)] = new_rec
    return _dump_json_or_jsonl(records, is_jsonl), {"key_rename": {"renamed": renamed[:10]}}


def _nested_inconsistency(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Wrap some values in arrays or objects."""
    records, is_jsonl = _parse_json_or_jsonl(content)
    if not records:
        return content, {"nested_inconsistency": False}
    targets = rng.choice(len(records), size=max(1, len(records) // 5), replace=False)
    for idx in targets:
        rec = records[int(idx)]
        if not isinstance(rec, dict):
            continue
        keys = list(rec.keys())
        if keys:
            k = py_rng.choice(keys)
            if py_rng.random() < 0.5:
                rec[k] = [rec[k]]
            else:
                rec[k] = {"value": rec[k]}
    return _dump_json_or_jsonl(records, is_jsonl), {"nested_inconsistency": {"count": len(targets)}}


def _missing_keys(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Remove random keys from some records."""
    records, is_jsonl = _parse_json_or_jsonl(content)
    if not records:
        return content, {"missing_keys": False}
    targets = rng.choice(len(records), size=max(1, len(records) // 5), replace=False)
    removed: list[str] = []
    for idx in targets:
        rec = records[int(idx)]
        if not isinstance(rec, dict):
            continue
        keys = list(rec.keys())
        if len(keys) > 1:
            k = py_rng.choice(keys)
            del rec[k]
            removed.append(k)
    return _dump_json_or_jsonl(records, is_jsonl), {"missing_keys": {"removed": removed[:10]}}


def _mixed_value_types(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Convert between string and numeric representations."""
    records, is_jsonl = _parse_json_or_jsonl(content)
    if not records:
        return content, {"mixed_value_types": False}
    changed = 0
    for idx in rng.choice(len(records), size=max(1, len(records) // 5), replace=False):
        rec = records[int(idx)]
        if not isinstance(rec, dict):
            continue
        for k in list(rec.keys()):
            v = rec[k]
            if isinstance(v, (int, float)) and py_rng.random() < 0.5:
                rec[k] = str(v)
                changed += 1
            elif isinstance(v, str):
                try:
                    rec[k] = int(v)
                    changed += 1
                except (ValueError, TypeError):
                    try:
                        rec[k] = float(v)
                        changed += 1
                    except (ValueError, TypeError):
                        pass
    return _dump_json_or_jsonl(records, is_jsonl), {"mixed_value_types": {"changed": changed}}


def _extra_wrapper(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Wrap the array in {"data": [...]}."""
    records, _is_jsonl = _parse_json_or_jsonl(content)
    wrapped = {"data": records}
    return json.dumps(wrapped, indent=2, default=str), {"extra_wrapper": True}


def _json_syntax_noise(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Inject trailing commas and extra whitespace (makes JSON technically invalid)."""
    # Add trailing commas before closing brackets
    result = re.sub(r'(\})\s*(\])', r'\1,\n\2', content, count=1)
    # Add extra whitespace around some colons
    lines = result.split("\n")
    targets = _pick_rows(lines, rng, 0.15) if len(lines) > 1 else []
    for i in targets:
        lines[i] = lines[i].replace(":", " :  ", 1)
    return "\n".join(lines), {"json_syntax_noise": True}


def _array_irregularity(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Convert some records to arrays (lose key information)."""
    records, is_jsonl = _parse_json_or_jsonl(content)
    if not records:
        return content, {"array_irregularity": False}
    targets = rng.choice(len(records), size=max(1, len(records) // 5), replace=False)
    for idx in targets:
        rec = records[int(idx)]
        if isinstance(rec, dict):
            records[int(idx)] = list(rec.values())
    return _dump_json_or_jsonl(records, is_jsonl), {"array_irregularity": {"count": len(targets)}}


# ====================================================================
# Excel corruptions (work on bytes via openpyxl)
# ====================================================================

def _merged_cells(content: bytes, rng: np.random.Generator, py_rng) -> tuple[bytes, dict]:
    """Merge random cell ranges."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    max_row = ws.max_row or 2
    max_col = ws.max_column or 2
    merges = 0
    for _ in range(min(3, max(1, max_row // 3))):
        r = py_rng.randint(2, max(2, max_row))
        c = py_rng.randint(1, max(1, max_col))
        try:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=min(c + 1, max_col))
            merges += 1
        except Exception:
            pass
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {"merged_cells": {"count": merges}}


def _multi_sheet_scatter(content: bytes, rng: np.random.Generator, py_rng) -> tuple[bytes, dict]:
    """Split data across multiple sheets."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), {"multi_sheet_scatter": False}
    header = rows[0]
    mid = len(rows) // 2
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(list(header))
    for row in rows[mid:]:
        ws2.append(list(row))
    # Remove moved rows from first sheet
    for i in range(ws.max_row, mid, -1):
        ws.delete_rows(i)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {"multi_sheet_scatter": {"sheets": 2}}


def _formula_artifacts(content: bytes, rng: np.random.Generator, py_rng) -> tuple[bytes, dict]:
    """Replace some cell values with formula-like strings."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    max_row = ws.max_row or 2
    max_col = ws.max_column or 2
    formulas = ["=SUM(A1:A10)", "=VLOOKUP(A1,B:C,2)", "=IF(A1>0,A1,0)", "=CONCAT(A1,B1)"]
    changed = 0
    n = max(1, max_row // 5)
    rows_to_hit = rng.choice(range(2, max(3, max_row + 1)), size=min(n, max_row - 1), replace=False)
    for r in rows_to_hit:
        c = py_rng.randint(1, max(1, max_col))
        ws.cell(row=int(r), column=c).value = py_rng.choice(formulas)
        changed += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {"formula_artifacts": {"changed": changed}}


def _header_misalignment(content: bytes, rng: np.random.Generator, py_rng) -> tuple[bytes, dict]:
    """Shift header down to row 2 or 3, leaving empty rows above."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    shift = py_rng.randint(1, 3)
    ws.insert_rows(1, amount=shift)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {"header_misalignment": {"shifted_by": shift}}


def _mixed_header_rows(content: bytes, rng: np.random.Generator, py_rng) -> tuple[bytes, dict]:
    """Insert extra header-like rows."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    # Insert a variation of the header at row 2
    alt_header = [str(h).upper() if h else h for h in header]
    ws.insert_rows(2)
    for c, val in enumerate(alt_header, 1):
        ws.cell(row=2, column=c).value = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {"mixed_header_rows": True}


# ====================================================================
# XML corruptions
# ====================================================================

def _inconsistent_tags(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Mix naming conventions in XML tags (camelCase, PascalCase, etc.)."""
    lines = content.split("\n")
    tag_re = re.compile(r"<(/?)(\w+)([ />])")
    changed_tags: list[str] = []
    for i, line in enumerate(lines):
        m = tag_re.search(line)
        if m and py_rng.random() < 0.2:
            tag = m.group(2)
            if tag not in ("data", "record", "xml"):
                new_tag = tag.upper() if py_rng.random() < 0.5 else tag.capitalize()
                lines[i] = line[:m.start(2)] + new_tag + line[m.end(2):]
                changed_tags.append(f"{tag}->{new_tag}")
    return "\n".join(lines), {"inconsistent_tags": {"changed": changed_tags[:10]}}


def _attributes_vs_elements(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Convert some child elements to parent attributes."""
    lines = content.split("\n")
    result: list[str] = []
    i = 0
    converted = 0
    while i < len(lines):
        line = lines[i]
        if "<record>" in line and py_rng.random() < 0.3:
            # Collect child elements until </record>
            attrs: list[str] = []
            children: list[str] = []
            i += 1
            while i < len(lines) and "</record>" not in lines[i]:
                m = re.match(r"\s*<(\w+)>(.*?)</\1>", lines[i])
                if m and py_rng.random() < 0.5:
                    attrs.append(f'{m.group(1)}="{m.group(2)}"')
                    converted += 1
                else:
                    children.append(lines[i])
                i += 1
            attr_str = " " + " ".join(attrs) if attrs else ""
            if children:
                result.append(f"  <record{attr_str}>")
                result.extend(children)
                result.append(lines[i] if i < len(lines) else "  </record>")
            else:
                result.append(f"  <record{attr_str} />")
                if i < len(lines):
                    pass  # skip </record> since we self-closed
        else:
            result.append(line)
        i += 1
    return "\n".join(result), {"attributes_vs_elements": {"converted": converted}}


def _namespace_mess(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Add xmlns namespace declarations."""
    namespaces = [
        'xmlns:ns1="http://example.com/ns1"',
        'xmlns:ns2="http://example.com/ns2"',
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
    ]
    chosen = py_rng.sample(namespaces, min(2, len(namespaces)))
    content = content.replace("<data>", f"<data {' '.join(chosen)}>", 1)
    # Prefix some tags
    content = re.sub(r"<record>", lambda m: "<ns1:record>" if py_rng.random() < 0.3 else m.group(), content)
    content = re.sub(r"</record>", lambda m: "</ns1:record>" if py_rng.random() < 0.3 else m.group(), content)
    return content, {"namespace_mess": {"added": chosen}}


def _missing_closing_tags(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Remove some closing tags to create malformed XML."""
    lines = content.split("\n")
    removed = 0
    for i in range(len(lines)):
        m = re.match(r"(\s*)</(\w+)>", lines[i])
        if m and m.group(2) not in ("data", "xml") and py_rng.random() < 0.15:
            lines[i] = ""
            removed += 1
    return "\n".join(lines), {"missing_closing_tags": {"removed": removed}}


# ====================================================================
# Fixed-width corruptions
# ====================================================================

def _column_misalignment(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Shift columns by adding/removing spaces."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.2)
    for i in targets:
        shift = py_rng.randint(1, 4)
        if py_rng.random() < 0.5:
            lines[i] = " " * shift + lines[i]
        else:
            lines[i] = lines[i][min(shift, len(lines[i])):]
    return "\n".join(lines), {"column_misalignment": {"rows_affected": len(targets)}}


def _inconsistent_padding(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Mix left and right padding."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.3)
    for i in targets:
        # Right-justify instead of left-justify by stripping and rjust
        stripped = lines[i].rstrip()
        orig_len = len(lines[i])
        if orig_len > len(stripped):
            lines[i] = stripped.rjust(orig_len)
    return "\n".join(lines), {"inconsistent_padding": {"rows_affected": len(targets)}}


def _shifted_fields(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Shift field boundaries by inserting/removing whitespace mid-field."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.15)
    for i in targets:
        line = lines[i]
        # Find a block of multiple spaces (field boundary) and shift it
        m = re.search(r"(\S)\s{2,}(\S)", line)
        if m:
            pos = m.start() + 1
            shift = py_rng.choice([-2, -1, 1, 2])
            spaces = line[pos:m.end() - 1]
            new_len = max(1, len(spaces) + shift)
            lines[i] = line[:pos] + " " * new_len + line[m.end() - 1:]
    return "\n".join(lines), {"shifted_fields": {"rows_affected": len(targets)}}


# ====================================================================
# HTML table corruptions
# ====================================================================

def _colspan_rowspan_chaos(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Add random colspan/rowspan attributes."""
    td_re = re.compile(r"<td>")
    positions = list(td_re.finditer(content))
    if not positions:
        return content, {"colspan_rowspan_chaos": False}
    n = max(1, len(positions) // 5)
    selected = py_rng.sample(list(range(len(positions))), min(n, len(positions)))
    # Build result from end to not invalidate positions
    result = content
    for idx in sorted(selected, reverse=True):
        m = positions[idx]
        attr = f'colspan="{py_rng.randint(1, 3)}"' if py_rng.random() < 0.5 else f'rowspan="{py_rng.randint(1, 3)}"'
        result = result[:m.start()] + f"<td {attr}>" + result[m.end():]
    return result, {"colspan_rowspan_chaos": {"modified": len(selected)}}


def _nested_tables(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Wrap some cell values in nested sub-tables."""
    td_val_re = re.compile(r"<td>(.*?)</td>")
    positions = list(td_val_re.finditer(content))
    if not positions:
        return content, {"nested_tables": False}
    n = max(1, len(positions) // 8)
    selected = py_rng.sample(list(range(len(positions))), min(n, len(positions)))
    result = content
    for idx in sorted(selected, reverse=True):
        m = positions[idx]
        val = m.group(1)
        nested = f"<td><table><tr><td>{val}</td></tr></table></td>"
        result = result[:m.start()] + nested + result[m.end():]
    return result, {"nested_tables": {"wrapped": len(selected)}}


def _missing_td(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Remove some <td> tags (leave content orphaned in <tr>)."""
    td_re = re.compile(r"<td>(.*?)</td>")
    positions = list(td_re.finditer(content))
    if not positions:
        return content, {"missing_td": False}
    n = max(1, len(positions) // 6)
    selected = py_rng.sample(list(range(len(positions))), min(n, len(positions)))
    result = content
    for idx in sorted(selected, reverse=True):
        m = positions[idx]
        result = result[:m.start()] + m.group(1) + result[m.end():]
    return result, {"missing_td": {"removed": len(selected)}}


def _hidden_data(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Add hidden rows with style='display:none'."""
    insert_before = content.rfind("</tbody>")
    if insert_before == -1:
        insert_before = content.rfind("</table>")
    if insert_before == -1:
        return content, {"hidden_data": False}
    hidden_rows = []
    for _ in range(py_rng.randint(1, 3)):
        hidden_rows.append(
            '<tr style="display:none"><td>HIDDEN</td><td>secret_data</td></tr>'
        )
    injection = "\n".join(hidden_rows)
    result = content[:insert_before] + injection + "\n" + content[insert_before:]
    return result, {"hidden_data": {"rows_added": len(hidden_rows)}}


# ====================================================================
# SQL dump corruptions
# ====================================================================

def _syntax_errors(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Remove semicolons from some statements."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.2)
    removed = 0
    for i in targets:
        if lines[i].endswith(";"):
            lines[i] = lines[i][:-1]
            removed += 1
    return "\n".join(lines), {"syntax_errors": {"semicolons_removed": removed}}


def _inconsistent_quoting(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Mix single quotes, double quotes, and backticks."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.25)
    for i in targets:
        if py_rng.random() < 0.5:
            lines[i] = lines[i].replace("'", '"')
        else:
            lines[i] = lines[i].replace("'", "`")
    return "\n".join(lines), {"inconsistent_quoting": {"rows_affected": len(targets)}}


def _mixed_null(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Replace NULL with various representations."""
    alts = ["null", "None", "''", "N/A"]
    lines = content.split("\n")
    changed = 0
    for i in range(len(lines)):
        if "NULL" in lines[i] and py_rng.random() < 0.5:
            lines[i] = lines[i].replace("NULL", py_rng.choice(alts), 1)
            changed += 1
    return "\n".join(lines), {"mixed_null": {"changed": changed}}


# ====================================================================
# YAML corruptions
# ====================================================================

def _indentation_errors(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Introduce wrong indentation levels."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.2)
    for i in targets:
        line = lines[i]
        stripped = line.lstrip()
        cur_indent = len(line) - len(stripped)
        new_indent = max(0, cur_indent + py_rng.choice([-1, 1, 3, -2]))
        lines[i] = " " * new_indent + stripped
    return "\n".join(lines), {"indentation_errors": {"rows_affected": len(targets)}}


def _mixed_tabs_spaces(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Replace some leading spaces with tabs."""
    lines = content.split("\n")
    targets = _pick_rows(lines, rng, 0.25)
    for i in targets:
        line = lines[i]
        stripped = line.lstrip()
        cur_indent = len(line) - len(stripped)
        if cur_indent > 0:
            lines[i] = "\t" * (cur_indent // 2 or 1) + stripped
    return "\n".join(lines), {"mixed_tabs_spaces": {"rows_affected": len(targets)}}


def _type_ambiguity(content: str, rng: np.random.Generator, py_rng) -> tuple[str, dict]:
    """Add string representations that look like booleans/numbers/null."""
    replacements = {
        "true": ['"true"', '"yes"', '"True"'],
        "false": ['"false"', '"no"', '"False"'],
        "null": ['"null"', '"None"', '""'],
    }
    lines = content.split("\n")
    changed = 0
    for i in range(len(lines)):
        for target, alts in replacements.items():
            if lines[i].rstrip().endswith(f": {target}") and py_rng.random() < 0.3:
                lines[i] = lines[i].replace(f": {target}", f": {py_rng.choice(alts)}")
                changed += 1
    # Also convert some numbers to strings
    num_re = re.compile(r": (\d+\.?\d*)$")
    for i in range(len(lines)):
        m = num_re.search(lines[i].rstrip())
        if m and py_rng.random() < 0.2:
            lines[i] = lines[i][:m.start(1)] + f'"{m.group(1)}"' + lines[i][m.end(1):]
            changed += 1
    return "\n".join(lines), {"type_ambiguity": {"changed": changed}}


# ---------------------------------------------------------------------------
# FORMAT_CORRUPTION_FNS registry
# ---------------------------------------------------------------------------

FORMAT_CORRUPTION_FNS: dict[str, _CorrFn] = {
    # CSV
    "mixed_delimiter": _mixed_delimiter,
    "bom_marker": _bom_marker,
    "mixed_encoding": _mixed_encoding,
    "mixed_line_endings": _mixed_line_endings,
    "quoted_inconsistency": _quoted_inconsistency,
    "header_duplicate": _header_duplicate,
    # TSV
    "tab_in_field": _tab_in_field,
    "mixed_tab_space": _mixed_tab_space,
    # JSON
    "key_rename": _key_rename,
    "nested_inconsistency": _nested_inconsistency,
    "missing_keys": _missing_keys,
    "mixed_value_types": _mixed_value_types,
    "extra_wrapper": _extra_wrapper,
    "json_syntax_noise": _json_syntax_noise,
    "array_irregularity": _array_irregularity,
    # Excel
    "merged_cells": _merged_cells,
    "multi_sheet_scatter": _multi_sheet_scatter,
    "formula_artifacts": _formula_artifacts,
    "header_misalignment": _header_misalignment,
    "mixed_header_rows": _mixed_header_rows,
    # XML
    "inconsistent_tags": _inconsistent_tags,
    "attributes_vs_elements": _attributes_vs_elements,
    "namespace_mess": _namespace_mess,
    "missing_closing_tags": _missing_closing_tags,
    # Fixed-width
    "column_misalignment": _column_misalignment,
    "inconsistent_padding": _inconsistent_padding,
    "shifted_fields": _shifted_fields,
    # HTML table
    "colspan_rowspan_chaos": _colspan_rowspan_chaos,
    "nested_tables": _nested_tables,
    "missing_td": _missing_td,
    "hidden_data": _hidden_data,
    # SQL dump
    "syntax_errors": _syntax_errors,
    "inconsistent_quoting": _inconsistent_quoting,
    "mixed_null": _mixed_null,
    # YAML
    "indentation_errors": _indentation_errors,
    "mixed_tabs_spaces": _mixed_tabs_spaces,
    "type_ambiguity": _type_ambiguity,
}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

_DIFFICULTY_COUNTS = {
    "easy": (1, 1),
    "medium": (0, 1),
    "hard": (2, 3),
}


def apply_format_corruptions(
    content: str | bytes,
    fmt: str,
    rng: np.random.Generator,
    py_rng,
    difficulty: str = "medium",
) -> tuple[str | bytes, list[dict]]:
    """Select and apply 1-3 format-specific corruptions based on *difficulty*.

    Returns (corrupted_content, list_of_metadata_dicts).
    """
    available = FORMAT_CORRUPTIONS.get(fmt, [])
    if not available:
        return content, []

    lo, hi = _DIFFICULTY_COUNTS.get(difficulty, (1, 2))
    n = py_rng.randint(lo, hi)
    n = min(n, len(available))

    chosen = py_rng.sample(available, n)
    all_meta: list[dict] = []

    for name in chosen:
        fn = FORMAT_CORRUPTION_FNS[name]
        content, meta = fn(content, rng, py_rng)
        all_meta.append(meta)

    return content, all_meta
